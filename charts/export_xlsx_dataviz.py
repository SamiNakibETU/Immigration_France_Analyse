"""
Exporte un classeur Excel (un onglet par data viz) pour Datawrapper, Flourish, Figma.
Exécution : python charts/export_xlsx_dataviz.py
Sortie : charts/output/terra_nova_migrations_dataviz.xlsx
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
CHARTS_OUT = ROOT / "charts" / "output"
SITE = ROOT / "site"
DATA_JSON = SITE / "data.json"

# Noms d’onglets ≤ 31 caractères (Excel)
SHEETS = {
    "readme": "00_README_et_sources",
    "f01": "01_FR_FX_DK_Danemark",
    "f02": "02_FR_FX_IT_Italie",
    "f03": "03_FR_FX_UK_2008plus",
    "f04": "04_FR_bleu_6pays_gris",
    "f05": "05_asile_lignes_5pays",
    "f06": "06_asile_barres_derniere",
    "f07": "07_solde_net_vs_asile",
    "f08": "08_snapshot_derniere_annee",
    "f09": "09_UE27_classement_2024",
    "f10": "10_rang_France_UE27",
    "f11": "11_ratio_asile_sur_solde_net",
    "f12": "12_quadri_FR_DE_IT_ES",
}


def _safe_sheet_name(name: str) -> str:
    return name[:31] if len(name) > 31 else name


def _read_csv_rows(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as f:
        return [list(row) for row in csv.reader(f)]


def _write_table(ws, rows: list[list[str]], start_row: int = 1, start_col: int = 1) -> int:
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row):
            c = ws.cell(row=start_row + ri, column=start_col + ci, value=val)
            if start_row + ri == 1 or (start_row == 1 and ri == 0):
                c.font = Font(bold=True)
    return start_row + len(rows)


def _load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def sheet_readme(wb: Workbook) -> None:
    ws = wb.create_sheet(_safe_sheet_name(SHEETS["readme"]), 0)
    lines = [
        ["Terra Nova — export données graphiques (Migrations)"],
        [""],
        ["Unités"],
        [
            "Solde migratoire net",
            "Pour 1 000 habitants — Eurostat demo_gind, indicateur CNMIGRATRT (taux brut + ajustement statistique).",
        ],
        [
            "Royaume-Uni (fig. 3, 8)",
            "Solde net de long terme ONS (pas Eurostat) ; dénominateur population documenté dans la méthodo.",
        ],
        [
            "Asile",
            "Premières demandes pour 1 000 habitants — migr_asyappctza (FRST) / demo_pjan.",
        ],
        [
            "FX (métropole)",
            "Souvent tronqué après ~2012 côté Eurostat pour CNMIGRATRT — ne pas prolonger artificiellement.",
        ],
        [""],
        ["Limites / angles non couverts ici"],
        [
            "Motifs de migration",
            "Pas de ventilation travail / famille / études dans ces séries (autres tables Eurostat ou données nationales).",
        ],
        [
            "Décisions d’asile",
            "Les graphiques utilisent des demandes initiales, pas le taux d’acceptation final.",
        ],
        [
            "Territoire France",
            "Pas de découpage régional (données nationales harmonisées seulement).",
        ],
        [
            "Rang UE (fig. 10)",
            "Basé sur les 27 États membres avec valeur pour l’année — pas le même échantillon si pays manquants.",
        ],
        [""],
        ["Fichier généré pour outils : Datawrapper, Flourish, Figma (reprise manuelle des colonnes)."],
        [""],
        ["Fichiers sources dans charts/output/ : voir noms des CSV alignés sur chaque onglet."],
    ]
    for i, row in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=row[0])
        if len(row) > 1:
            ws.cell(row=i, column=2, value=row[1])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 88
    ws["A1"].font = Font(bold=True, size=14)


def sheet_from_csv(wb: Workbook, key: str, csv_name: str, title_note: str = "") -> None:
    path = CHARTS_OUT / csv_name
    rows = _read_csv_rows(path)
    ws = wb.create_sheet(_safe_sheet_name(SHEETS[key]))
    r = 1
    if title_note:
        ws.cell(row=r, column=1, value=title_note)
        ws.cell(row=r, column=1).font = Font(italic=True, size=9)
        r += 2
    if rows:
        _write_table(ws, rows, start_row=r)
    else:
        ws.cell(row=r, column=1, value="(fichier absent — lancez plot_publication.py)")


def sheet_fig04(wb: Workbook, mig_all: list[list[str]]) -> None:
    """Deux blocs : haut FR+FX, bas FR + 6 pays (même logique que le site)."""
    ws = wb.create_sheet(_safe_sheet_name(SHEETS["f04"]))
    if not mig_all:
        ws.cell(row=1, column=1, value="Pas de données — cnmigratrt_2005_2024_tous_pays.csv manquant.")
        return
    header = mig_all[0]
    idx = {h: i for i, h in enumerate(header)}
    r = 1
    ws.cell(row=r, column=1, value="Bloc A — Panneau haut (France FR + FX)")
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    cols_a = ["year", "FR", "FX"]
    ha = [c for c in cols_a if c in idx]
    row_h = [ha]
    for row in mig_all[1:]:
        row_h.append([row[idx[c]] if c in idx else "" for c in ha])
    _write_table(ws, row_h, start_row=r)
    r += len(row_h) + 2

    ws.cell(row=r, column=1, value="Bloc B — Panneau bas (FR bleu + DE ES SE NL IT DK en gris sur le graphique)")
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    cols_b = ["year", "FR", "DE", "ES", "SE", "NL", "IT", "DK"]
    hb = [c for c in cols_b if c in idx]
    row_b = [hb]
    for row in mig_all[1:]:
        row_b.append([row[idx[c]] if c in idx else "" for c in hb])
    _write_table(ws, row_b, start_row=r)


def _mig_all_rows() -> list[list[str]]:
    return _read_csv_rows(CHARTS_OUT / "cnmigratrt_2005_2024_tous_pays.csv")


def sheet_fig01_03_wide(wb: Workbook) -> None:
    sel = _read_csv_rows(CHARTS_OUT / "cnmigratrt_2005_2024_selection.csv")
    if len(sel) < 2:
        return
    header = sel[0]
    idx = {h: i for i, h in enumerate(header)}

    def write_fig(key: str, cols: list[str], note: str) -> None:
        ws = wb.create_sheet(_safe_sheet_name(SHEETS[key]))
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, size=9)
        h = [c for c in cols if c in idx]
        data = [h]
        for row in sel[1:]:
            data.append([row[idx[c]] if c in idx else "" for c in h])
        _write_table(ws, data, start_row=3)

    write_fig(
        "f01",
        ["year", "FR", "FX", "DK"],
        "Fig. 1 — Site : deux panneaux (haut FR+FX, bas DK). Colonnes pour série temporelle ou format long.",
    )
    write_fig(
        "f02",
        ["year", "FR", "FX", "IT"],
        "Fig. 2 — France, métropole FX, Italie.",
    )
    write_fig(
        "f03",
        ["year", "FR", "FX", "UK"],
        "Fig. 3 — Filtrer year >= 2008 pour l’affichage web.",
    )


def sheet_from_json_table(
    wb: Workbook,
    key: str,
    data: list | None,
    columns: list[str],
    note: str = "",
) -> None:
    ws = wb.create_sheet(_safe_sheet_name(SHEETS[key]))
    r = 1
    if note:
        ws.cell(row=r, column=1, value=note)
        r += 2
    if not data:
        ws.cell(row=r, column=1, value="(vide — exécutez build_data.py)")
        return
    ws.append(columns)
    for row in data:
        ws.append([row.get(c, "") for c in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)


def main() -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)

    # Fig 1–3 : largeur depuis sélection
    sheet_fig01_03_wide(wb)

    # Fig 4
    sheet_fig04(wb, _mig_all_rows())

    # Fig 5–6, 9–12 : CSV
    sheet_from_csv(
        wb,
        "f05",
        "asile_premieres_demandes_pour_1000.csv",
        "Fig. 5 — Colonnes pays (hors année) = séries pour graphiques en lignes.",
    )
    sheet_from_csv(wb, "f09", "eu27_cnmigratrt_2024_from_api.csv", "Fig. 9 — Classement UE-27, millésime 2024.")
    sheet_from_csv(wb, "f10", "analyse_rang_france_ue27.csv", "Fig. 10 — Rang France parmi UE-27 (1 = solde net le plus élevé).")
    sheet_from_csv(wb, "f11", "analyse_ratio_asile_solde_net.csv", "Fig. 11 — Ratio asile/solde net si solde net > 0.")
    sheet_from_csv(wb, "f12", "analyse_quadri_fr_de_it_es.csv", "Fig. 12 — Quadri grandes économies.")

    data = _load_json(DATA_JSON)

    # Fig 6 — barres asile dernière année
    sheet_from_json_table(
        wb,
        "f06",
        data.get("asylumBars2024"),
        ["code", "label", "value"],
        "Fig. 6 — Dernière année disponible dans les données (colonne label = pays + année).",
    )

    # Fig 7 — double barres
    dual = data.get("dualNetAsylum2024") or {}
    ws = wb.create_sheet(_safe_sheet_name(SHEETS["f07"]))
    ws.cell(row=1, column=1, value="Fig. 7 — Gauche : solde net (dernière année par pays). Droite : asile.")
    ws.cell(row=1, column=1).font = Font(bold=True)
    r = 3
    ws.cell(row=r, column=1, value="Solde migratoire net (CNMIGRATRT)")
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    net = dual.get("net") or []
    if net:
        ws.cell(row=r, column=1, value="code")
        ws.cell(row=r, column=2, value="label")
        ws.cell(row=r, column=3, value="year")
        ws.cell(row=r, column=4, value="value")
        r += 1
        for row in net:
            ws.cell(row=r, column=1, value=row.get("code", ""))
            ws.cell(row=r, column=2, value=row.get("label", ""))
            ws.cell(row=r, column=3, value=row.get("year", ""))
            ws.cell(row=r, column=4, value=row.get("value", ""))
            r += 1
    r += 2
    ws.cell(row=r, column=1, value="Premières demandes d’asile (pour 1 000 hab.)")
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    asy = dual.get("asylum") or []
    if asy:
        ws.cell(row=r, column=1, value="code")
        ws.cell(row=r, column=2, value="label")
        ws.cell(row=r, column=3, value="year")
        ws.cell(row=r, column=4, value="value")
        r += 1
        for row in asy:
            ws.cell(row=r, column=1, value=row.get("code", ""))
            ws.cell(row=r, column=2, value=row.get("label", ""))
            ws.cell(row=r, column=3, value=row.get("year", ""))
            ws.cell(row=r, column=4, value=row.get("value", ""))
            r += 1

    # Fig 8 — snapshot
    sheet_from_json_table(
        wb,
        "f08",
        data.get("snapshot"),
        ["code", "label", "year", "value", "barKey", "yLabel"],
        "Fig. 8 — Dernière année par série (solde net).",
    )

    out_path = CHARTS_OUT / "terra_nova_migrations_dataviz.xlsx"
    wb.save(out_path)
    print(f"OK — {out_path}")
    return out_path


if __name__ == "__main__":
    main()

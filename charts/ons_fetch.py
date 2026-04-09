"""
Télécharge le classeur ONS « Long-term international migration flows, provisional »
et extrait le solde net (All Nationalities) pour chaque **Year Ending December** (YE Dec).

Dernière édition suivie par défaut : year ending June 2025 (ltimnov25.xlsx) — révise les millésimes récents.
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

import openpyxl

ONS_USER_AGENT = (
    "Mozilla/5.0 (compatible; TerraNova-MigrationResearch/1.0; "
    "+https://github.com) Python urllib"
)

# Édition la plus récente sur la page dataset ONS (révisions possibles vs édition « Dec 2024 » figée)
LTIM_LATEST_XLSX_URL = (
    "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/populationandmigration/"
    "internationalmigration/datasets/longterminternationalimmigrationemigrationandnetmigrationflowsprovisional/"
    "yearendingjune2025/ltimnov25.xlsx"
)


def _download(url: str, timeout: int = 120) -> bytes:
    req = Request(url, headers={"User-Agent": ONS_USER_AGENT})
    with urlopen(req, timeout=timeout) as resp:
        return resp.read()


def _open_workbook(data: bytes) -> Any:
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def _main_table_sheet(wb: Any) -> Any:
    for name in ("Table 1", "1"):
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        if name.startswith("Table "):
            return wb[name]
    raise ValueError(f"Aucune feuille Table 1 / 1 dans le classeur ONS : {wb.sheetnames[:8]}")


def parse_net_migration_ye_december(sheet: Any) -> dict[int, int]:
    """
    Lignes : Flow = 'Net migration', Period contient 'YE Dec yy', valeur All Nationalities en colonne C (index 2).
    """
    out: dict[int, int] = {}
    dec_re = re.compile(r"YE\s+Dec\s+(\d{2})\b", re.IGNORECASE)
    for row in sheet.iter_rows(values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        flow = str(row[0]).strip()
        if flow != "Net migration":
            continue
        period = str(row[1]).strip()
        m = dec_re.search(period)
        if not m:
            continue
        yy = int(m.group(1))
        year = 2000 + yy if yy < 70 else 1900 + yy
        val = row[2]
        if isinstance(val, bool):
            continue
        if isinstance(val, (int, float)):
            out[year] = int(val)
    return out


def refresh_ons_ltim(
    cache_root: Path,
    url: str | None = None,
) -> dict[str, Any]:
    """
    Télécharge le XLSX, enregistre sous cache_root/ons/, écrit ons_ltim_net_ye_dec.json.
    Retourne le dict métadonnées + série parsée.
    """
    url = url or LTIM_LATEST_XLSX_URL
    ons_dir = cache_root / "ons"
    ons_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = ons_dir / "ltim_latest.xlsx"
    meta_path = ons_dir / "ons_ltim_net_ye_dec.json"

    raw = _download(url)
    xlsx_path.write_bytes(raw)
    wb = _open_workbook(raw)
    try:
        ws = _main_table_sheet(wb)
        parsed = parse_net_migration_ye_december(ws)
    finally:
        wb.close()

    meta: dict[str, Any] = {
        "source_url": url,
        "downloaded_utc": datetime.now(timezone.utc).isoformat(),
        "xlsx_file": str(xlsx_path.relative_to(cache_root)),
        "ye_december_net_all_nationalities_persons": {str(y): v for y, v in sorted(parsed.items())},
        "years_parsed": len(parsed),
    }
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    return meta


def load_parsed_ye_dec_from_cache(cache_root: Path) -> dict[int, int] | None:
    p = cache_root / "ons" / "ons_ltim_net_ye_dec.json"
    if not p.exists():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    block = data.get("ye_december_net_all_nationalities_persons") or {}
    out: dict[int, int] = {}
    for k, v in block.items():
        try:
            out[int(k)] = int(v)
        except (TypeError, ValueError):
            continue
    return out if out else None
